# -*- coding: utf-8 -*-
"""一次性迁移脚本（2026-08-15 家具族化）：把 Excel/家具表.xlsx 拆成家具族表 + 瘦身后的家具表。

做三件事：
  ① 按 id 前缀（`<族>_<编号>`）分族，**逐族核对族级列是否真的族内一致**——不一致就报错退出，
     绝不"取第一行的值"糊过去（那会静默改掉某些家具的占格/装饰分）。
  ② 生成 Excel/家具族表.xlsx（族级列）。
  ③ 把 Excel/家具表.xlsx 改写成只剩变体列 + 新增「族id」列。

跑完之后再跑 export_config.bat 出 CSV。**这个脚本只用跑一次**，跑完可以留在仓库里当迁移记录。
原表会先备份成 Excel/家具表.xlsx.bak（已存在则不覆盖备份）。
"""
import os
import re
import shutil
import sys
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
FURNITURE_XLSX = os.path.join(ROOT, "Excel", "家具表.xlsx")
FAMILY_XLSX = os.path.join(ROOT, "Excel", "家具族表.xlsx")
BACKUP = FURNITURE_XLSX + ".bak"

FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")

# 族级列：同族必然相同 → 搬去族表（列名, 族表字段名）
FAMILY_COLUMNS = [
    ("分类", "category"),
    ("描述", "description"),
    ("表面类型", "surfaces"),
    ("可叠放", "stackable"),
    ("占格列", "cols"),
    ("占格行", "rows"),
    ("装饰分", "decorationScore"),
    ("拿起音效", "pickupSound"),
    ("放下音效", "putdownSound"),
    ("桌面格启用", "tableSurface.enabled"),
    ("桌面格列数", "tableSurface.cols"),
    ("桌面格宽", "tableSurface.cellWidth"),
    ("桌面格高", "tableSurface.cellHeight"),
    ("桌面格偏移X", "tableSurface.offsetX"),
    ("桌面高度", "tableSurface.surfaceHeight"),
]

# 变体列：逐个不同 → 留在家具表（列名, 字段名）
VARIANT_COLUMNS = [
    ("id", "id"),
    ("英文索引", "nameKey"),
    ("显示名", "displayName"),
    ("族id", "familyId"),
    ("显示宽", "displayWidth"),
    ("显示高", "displayHeight"),
    ("精灵图", "sprite"),
    ("色值", "swatchColor"),
]


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def family_id_of(furniture_id):
    """族 id = 最后一个 `_` 之前的部分（现有 121 行 id 已规整，零手填，见文档 §6 待确认 #4）。"""
    return furniture_id.rsplit("_", 1)[0] if "_" in furniture_id else furniture_id


def family_display_name(variant_display_name):
    """族显示名 = 变体名去掉「·NN」编号后缀（`单人沙发·02` → `单人沙发`）。"""
    return variant_display_name.split("·", 1)[0] if "·" in variant_display_name else variant_display_name


def read_rows():
    wb = openpyxl.load_workbook(FURNITURE_XLSX)
    ws = wb["家具"]
    raw = list(ws.iter_rows(values_only=True))
    header = [cell_text(c) for c in raw[0]]
    index = {name: header.index(name) for name in header if name}

    rows = []
    for excel_row, values in enumerate(raw[1:], start=2):
        record = {name: cell_text(values[i]) if i < len(values) else "" for name, i in index.items()}
        if not any(record.values()):
            continue
        if excel_row == 2 and all(FIELD_NAME.match(v) for v in record.values() if v):
            continue  # 第 2 行 = Unity 字段名参考行
        if not record.get("id"):
            print(f"[WARN] 家具表 Excel 第 {excel_row} 行缺 id，已跳过")
            continue
        record["_excel_row"] = excel_row
        rows.append(record)
    wb.close()
    return rows


def group_families(rows):
    families = OrderedDict()
    for record in rows:
        families.setdefault(family_id_of(record["id"]), []).append(record)
    return families


def verify_family_columns(families):
    """族级列必须族内一致。这是整个族化改造的前提，不成立就说明族划错了或列选错了。"""
    problems = []
    for family_id, members in families.items():
        for column, _ in FAMILY_COLUMNS:
            values = {m.get(column, "") for m in members}
            if len(values) > 1:
                sample = ", ".join(f"{m['id']}={m.get(column, '')!r}" for m in members[:4])
                problems.append(f"  族「{family_id}」的「{column}」族内不一致：{sorted(values)}（{sample} …）")
    if problems:
        print("[ERROR] 以下列在族内并不一致，不能搬进族表——请先确认族划分或把该列留在家具表：")
        print("\n".join(problems))
        sys.exit(1)
    print(f"[OK] 族级列族内一致性校验通过：{len(families)} 个族 × {len(FAMILY_COLUMNS)} 列")


def write_family_xlsx(families):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "家具族"
    header = ["族id", "族显示名"] + [name for name, _ in FAMILY_COLUMNS]
    fields = ["familyId", "displayName"] + [field for _, field in FAMILY_COLUMNS]
    ws.append(header)
    ws.append(fields)
    for cell in ws[2]:
        cell.font = Font(italic=True, color="808080")  # 第 2 行 = 字段名参考行（灰色斜体，与其它表一致）
    for family_id, members in families.items():
        first = members[0]
        ws.append([family_id, family_display_name(first.get("显示名", ""))] +
                  [first.get(name, "") for name, _ in FAMILY_COLUMNS])
    ws.freeze_panes = "A3"
    for i, name in enumerate(header, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(10, len(name) * 2 + 4)
    wb.save(FAMILY_XLSX)
    print(f"[OK] 已生成 {os.path.abspath(FAMILY_XLSX)}：{len(families)} 个族")


def rewrite_furniture_xlsx(rows):
    if not os.path.exists(BACKUP):
        shutil.copyfile(FURNITURE_XLSX, BACKUP)
        print(f"[OK] 原表已备份到 {os.path.abspath(BACKUP)}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "家具"
    ws.append([name for name, _ in VARIANT_COLUMNS])
    ws.append([field for _, field in VARIANT_COLUMNS])
    for cell in ws[2]:
        cell.font = Font(italic=True, color="808080")
    for record in rows:
        record["族id"] = family_id_of(record["id"])
        ws.append([record.get(name, "") for name, _ in VARIANT_COLUMNS])
    ws.freeze_panes = "A3"
    widths = {"id": 24, "英文索引": 24, "显示名": 18, "族id": 22, "显示宽": 9, "显示高": 9, "精灵图": 52, "色值": 10}
    for i, (name, _) in enumerate(VARIANT_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(name, 14)
    wb.save(FURNITURE_XLSX)
    print(f"[OK] 已改写 {os.path.abspath(FURNITURE_XLSX)}：{len(rows)} 行 × {len(VARIANT_COLUMNS)} 列")


def main():
    if not os.path.exists(FURNITURE_XLSX):
        print(f"[ERROR] Excel not found: {os.path.abspath(FURNITURE_XLSX)}")
        sys.exit(1)
    rows = read_rows()
    print(f"[INFO] 读到 {len(rows)} 行家具")
    families = group_families(rows)
    verify_family_columns(families)
    write_family_xlsx(families)
    rewrite_furniture_xlsx(rows)
    print("[DONE] 迁移完成。下一步：跑 Tools/导表/export_config.bat 出 CSV。")


if __name__ == "__main__":
    main()
