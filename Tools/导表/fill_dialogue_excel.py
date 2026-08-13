# -*- coding: utf-8 -*-
"""填写对话表.xlsx 的 demo 文案，然后跑 export_config.bat 正常导出。"""
import os, openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

ROOT       = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL_PATH = os.path.join(ROOT, "Excel", "对话表.xlsx")
FONT_NAME  = "微软雅黑"

THIN = Side(style="thin", color="AAAAAA")
BD   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def cell_style(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row, c)
        cell.border  = BD
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font    = Font(name=FONT_NAME)

# ── 对话数据 ────────────────────────────────────────────────────────────────
# (对话组ID, 备注, 文件夹, 步骤, 类型, 说话人, 表情, 文本, 动作, 动作参数, 跳转, 跳转目标组, 选项条件)
CONTENT = [
    # 乌鸦
    ("crow_firstMeeting_01","乌鸦·初次见面","通用",1,"台词","访客","平静","……我听说这里很安静。","","","","",""),
    ("","","","",2,"选项","","","请进。","接待","","结束","","有空房"),
    ("","","","",2,"选项","","","今天客满了。","拒绝","","结束","",""),
    ("crow_serviceStart_01","乌鸦·开始服务","通用",1,"台词","访客","平静","房间有窗户吗？面朝海的那种。","","","","",""),
    ("crow_serviceCheck_01","乌鸦·服务中","通用",1,"台词","访客","平静","窗框的风声……我觉得是什么东西在应答我。","","","","",""),
    ("crow_serviceCheck_02","乌鸦·完成","通用",1,"台词","访客","高兴","昨晚我唱了一句，它应了一句。不是重复——是答案。","","","","",""),
    ("","","","",2,"选项","","","听起来不错。","完成需求","","结束","","房间有家具"),
    ("","","","",2,"选项","","","（继续打理房间）","","","结束","",""),
    ("crow_rejected_01","乌鸦·拒绝","通用",1,"台词","访客","平静","……好，我再去找找。","","","","",""),
    ("crow_doneMismatch_01","乌鸦·不对味","通用",1,"台词","访客","失望","……不是这个。说不清哪里不对。","","","","",""),
    ("crow_donePlain_01","乌鸦·一般","通用",1,"台词","访客","平静","……还行，比以前住的好一点。","","","","",""),
    ("crow_doneSatisfied_01","乌鸦·满意","通用",1,"台词","访客","高兴","谢谢你。……这句话我不常说。","","","","",""),
    ("crow_donePerfect_01","乌鸦·完美","通用",1,"台词","访客","高兴","它把我每一句都接住了。没有一句落地。","","","","",""),
    ("","","","",2,"台词","访客","平静","……我三年没说过不错了。","","","","",""),
    ("crow_wanderChat_01","乌鸦·闲逛","通用",1,"台词","访客","平静","（他对着海，嘴唇微微动了一下）","","","","",""),
    # 狐狸
    ("fox_firstMeeting_01","狐狸·初次见面","通用",1,"台词","访客","平静","这地方没有评分？没有评分的地方要么很好要么很烂。","","","","",""),
    ("","","","",2,"选项","","","那进来看看才知道。","接待","","结束","","有空房"),
    ("","","","",2,"选项","","","今天客满了。","拒绝","","结束","",""),
    ("fox_serviceStart_01","狐狸·开始服务","通用",1,"台词","访客","平静","采光、隔音……还有一个我说不清楚的东西。你布置就好，我会告诉你哪里不对。","","","","",""),
    ("fox_serviceCheck_01","狐狸·服务中","通用",1,"台词","访客","平静","还差一点。等我知道了再告诉你。","","","","",""),
    ("fox_serviceCheck_02","狐狸·完成","通用",1,"台词","访客","高兴","我算过了——这是我住过最划算的一次。","","","","",""),
    ("","","","",2,"选项","","","那就好。","完成需求","","结束","","房间有家具"),
    ("","","","",2,"选项","","","（继续完善房间）","","","结束","",""),
    ("fox_rejected_01","狐狸·拒绝","通用",1,"台词","访客","平静","理解。下次再说。","","","","",""),
    ("fox_doneMismatch_01","狐狸·不对味","通用",1,"台词","访客","失望","……不是我要的。也不是你的问题。","","","","",""),
    ("fox_donePlain_01","狐狸·一般","通用",1,"台词","访客","平静","中规中矩。比预期差一点，但也没差太多。","","","","",""),
    ("fox_doneSatisfied_01","狐狸·满意","通用",1,"台词","访客","高兴","不错。……我很少这样说。","","","","",""),
    ("fox_donePerfect_01","狐狸·完美","通用",1,"台词","访客","高兴","我住过的地方有一百多个。这是第一个让我不想算账的。","","","","",""),
    ("fox_wanderChat_01","狐狸·闲逛","通用",1,"台词","访客","平静","（他看着窗外，眼神很远）","","","","",""),
    # 刺猬
    ("hedgehog_firstMeeting_01","刺猬·初次见面","通用",1,"台词","访客","平静","这里……大家都很安静吗？","","","","",""),
    ("","","","",2,"选项","","","是的，欢迎进来。","接待","","结束","","有空房"),
    ("","","","",2,"选项","","","今天客满了。","拒绝","","结束","",""),
    ("hedgehog_serviceStart_01","刺猬·开始服务","通用",1,"台词","访客","平静","我想要一个柔软的房间。……我也不知道什么叫柔软的房间。","","","","",""),
    ("hedgehog_serviceCheck_01","刺猬·服务中","通用",1,"台词","访客","平静","我……没有扎到什么东西吧？","","","","",""),
    ("hedgehog_serviceCheck_02","刺猬·完成","通用",1,"台词","访客","高兴","那个沙发……它抱住我了。我以为我会被弹走。","","","","",""),
    ("","","","",2,"选项","","","没有弹走你。","完成需求","","结束","","房间有家具"),
    ("","","","",2,"选项","","","（继续完善房间）","","","结束","",""),
    ("hedgehog_rejected_01","刺猬·拒绝","通用",1,"台词","访客","平静","……没关系，我习惯了。","","","","",""),
    ("hedgehog_doneMismatch_01","刺猬·不对味","通用",1,"台词","访客","失望","……还是扎到了什么。不是你的错，是我太尖了。","","","","",""),
    ("hedgehog_donePlain_01","刺猬·一般","通用",1,"台词","访客","平静","……还好，比我预计的好一点。","","","","",""),
    ("hedgehog_doneSatisfied_01","刺猬·满意","通用",1,"台词","访客","高兴","我睡得很好。……好久没说过这句话了。","","","","",""),
    ("hedgehog_donePerfect_01","刺猬·完美","通用",1,"台词","访客","高兴","我的刺只有害怕的时候才会硬。昨晚我一根都没硬。","","","","",""),
    ("hedgehog_wanderChat_01","刺猬·闲逛","通用",1,"台词","访客","平静","（她轻手轻脚地走着，生怕碰到什么）","","","","",""),
    # 兔子
    ("rabbit_firstMeeting_01","兔子·初次见面","通用",1,"台词","访客","高兴","你好！我想住一住，如果可以的话！","","","","",""),
    ("","","","",2,"选项","","","请进！","接待","","结束","","有空房"),
    ("","","","",2,"选项","","","今天客满了。","拒绝","","结束","",""),
    ("rabbit_serviceStart_01","兔子·开始服务","通用",1,"台词","访客","高兴","怎么布置都可以！……没有人问过我想要什么。","","","","",""),
    ("rabbit_serviceCheck_01","兔子·服务中","通用",1,"台词","访客","困惑","有什么可以帮忙的吗？……哦，那我就休息一下？","","","","",""),
    ("rabbit_serviceCheck_02","兔子·完成","通用",1,"台词","访客","高兴","外面很吵的时候风铃不动。夜深了，它自己晃起来了。","","","","",""),
    ("","","","",2,"选项","","","好好休息。","完成需求","","结束","","房间有家具"),
    ("","","","",2,"选项","","","（继续完善房间）","","","结束","",""),
    ("rabbit_rejected_01","兔子·拒绝","通用",1,"台词","访客","高兴","没关系的！完全没关系！……（耳朵垂了下来）","","","","",""),
    ("rabbit_doneMismatch_01","兔子·不对味","通用",1,"台词","访客","高兴","挺好的！……对不起，我不知道怎么说不好。","","","","",""),
    ("rabbit_donePlain_01","兔子·一般","通用",1,"台词","访客","高兴","谢谢你！……（低头想了一下）确实不错。","","","","",""),
    ("rabbit_doneSatisfied_01","兔子·满意","通用",1,"台词","访客","平静","你第一个问过我想要什么。我一直在想这个问题。","","","","",""),
    ("rabbit_donePerfect_01","兔子·完美","通用",1,"台词","访客","高兴","我昨晚听见自己的脚步声了。以前走路很轻，怕打扰别人。昨晚我就正常走路了。","","","","",""),
    ("rabbit_wanderChat_01","兔子·闲逛","通用",1,"台词","访客","高兴","（她路过，想说什么又憋回去了）","","","","",""),
]

# (对话组ID, 文件夹, 种族, 触发分类, 权重, 进入条件)
POOL = [
    ("crow_firstMeeting_01",   "通用","crow",     "firstMeeting",  1,""),
    ("crow_serviceStart_01",   "通用","crow",     "serviceStart",  1,""),
    ("crow_serviceCheck_01",   "通用","crow",     "serviceCheck",  1,""),
    ("crow_serviceCheck_02",   "通用","crow",     "serviceCheck",  1,""),
    ("crow_rejected_01",       "通用","crow",     "rejected",      1,""),
    ("crow_doneMismatch_01",   "通用","crow",     "doneMismatch",  1,""),
    ("crow_donePlain_01",      "通用","crow",     "donePlain",     1,""),
    ("crow_doneSatisfied_01",  "通用","crow",     "doneSatisfied", 1,""),
    ("crow_donePerfect_01",    "通用","crow",     "donePerfect",   1,""),
    ("crow_wanderChat_01",     "通用","crow",     "wanderChat",    1,""),
    ("fox_firstMeeting_01",    "通用","fox",      "firstMeeting",  1,""),
    ("fox_serviceStart_01",    "通用","fox",      "serviceStart",  1,""),
    ("fox_serviceCheck_01",    "通用","fox",      "serviceCheck",  1,""),
    ("fox_serviceCheck_02",    "通用","fox",      "serviceCheck",  1,""),
    ("fox_rejected_01",        "通用","fox",      "rejected",      1,""),
    ("fox_doneMismatch_01",    "通用","fox",      "doneMismatch",  1,""),
    ("fox_donePlain_01",       "通用","fox",      "donePlain",     1,""),
    ("fox_doneSatisfied_01",   "通用","fox",      "doneSatisfied", 1,""),
    ("fox_donePerfect_01",     "通用","fox",      "donePerfect",   1,""),
    ("fox_wanderChat_01",      "通用","fox",      "wanderChat",    1,""),
    ("hedgehog_firstMeeting_01","通用","hedgehog","firstMeeting",  1,""),
    ("hedgehog_serviceStart_01","通用","hedgehog","serviceStart",  1,""),
    ("hedgehog_serviceCheck_01","通用","hedgehog","serviceCheck",  1,""),
    ("hedgehog_serviceCheck_02","通用","hedgehog","serviceCheck",  1,""),
    ("hedgehog_rejected_01",   "通用","hedgehog","rejected",       1,""),
    ("hedgehog_doneMismatch_01","通用","hedgehog","doneMismatch",  1,""),
    ("hedgehog_donePlain_01",  "通用","hedgehog","donePlain",      1,""),
    ("hedgehog_doneSatisfied_01","通用","hedgehog","doneSatisfied",1,""),
    ("hedgehog_donePerfect_01","通用","hedgehog","donePerfect",    1,""),
    ("hedgehog_wanderChat_01", "通用","hedgehog","wanderChat",     1,""),
    ("rabbit_firstMeeting_01", "通用","rabbit",  "firstMeeting",  1,""),
    ("rabbit_serviceStart_01", "通用","rabbit",  "serviceStart",  1,""),
    ("rabbit_serviceCheck_01", "通用","rabbit",  "serviceCheck",  1,""),
    ("rabbit_serviceCheck_02", "通用","rabbit",  "serviceCheck",  1,""),
    ("rabbit_rejected_01",     "通用","rabbit",  "rejected",       1,""),
    ("rabbit_doneMismatch_01", "通用","rabbit",  "doneMismatch",  1,""),
    ("rabbit_donePlain_01",    "通用","rabbit",  "donePlain",      1,""),
    ("rabbit_doneSatisfied_01","通用","rabbit",  "doneSatisfied", 1,""),
    ("rabbit_donePerfect_01",  "通用","rabbit",  "donePerfect",   1,""),
    ("rabbit_wanderChat_01",   "通用","rabbit",  "wanderChat",    1,""),
]

# ── 写入 Excel ───────────────────────────────────────────────────────────────
if not os.path.exists(EXCEL_PATH):
    print(f"[ERROR] 找不到 {EXCEL_PATH}")
    print("请先在 Unity 中运行 init_dialogue_excel.py 生成对话表.xlsx")
    exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH)

# Sheet 1 对话内容
ws1 = wb["对话内容"]
# 清除旧数据行（保留第1行表头和第2行注释）
for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row):
    for cell in row:
        cell.value = None

r = 3
for vals in CONTENT:
    for c, v in enumerate(vals, 1):
        ws1.cell(r, c, v if v != "" else None)
    cell_style(ws1, r, 13)
    r += 1

# Sheet 2 对话池
ws2 = wb["对话池"]
for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row):
    for cell in row:
        cell.value = None

r2 = 3
for vals in POOL:
    for c, v in enumerate(vals, 1):
        ws2.cell(r2, c, v if v != "" else None)
    cell_style(ws2, r2, 6)
    r2 += 1

wb.save(EXCEL_PATH)
print(f"[OK] 已写入 {EXCEL_PATH}")
print(f"     对话内容: {len(CONTENT)} 行")
print(f"     对话池:   {len(POOL)} 行")
print()
print("现在双击 Tools/导表/export_config.bat 导出 CSV。")
