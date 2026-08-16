# -*- coding: utf-8 -*-
"""生成 Excel/立绘表.xlsx 模板，并把工程里现有的立绘素材铺成起始内容。

用法：
    python Tools/导表/make_portrait_template.py          # 文件已存在时拒绝覆盖
    python Tools/导表/make_portrait_template.py --force  # 强制重建（会丢掉现有配置！）

生成之后正常流程就是「改 xlsx → 双击 export_config.bat → 切回 Unity」，本脚本不再需要。

这张表只有三列，刻意保持朴素：

    立绘ID        资源路径                        备注
    rabbit_平静   OutGameUI/Portraits/rabbit      兔族默认脸

立绘ID 是**主键**，对话表第二页的「立绘ID」列、访客种族表的「默认立绘ID」列引用它。
命名不强制规范——2026-08-14 撤掉表情枚举就是因为美术的差分归不了类，这里再立规矩等于白撤。
唯一的硬约束是 ID 不能带逗号（CSV 分隔符）。起始内容用 `<raceId>_平静` 只是为了跟
VisitorConfigSetupUtility 里的默认值对上，改成什么都行，改完记得同步种族表那一列。

资源路径写 **Resources 相对路径、不带扩展名**（如 OutGameUI/Portraits/rabbit）——
立绘走 Resources.Load<Texture2D> 取图，所以素材必须待在某个 Resources 目录下
（美术交付件躺在 Assets/Arts/对话大头，那儿不是 Resources，先用 Tools/install_visitor_art.py 拷进来）。
"""
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL_PATH = os.path.join(ROOT, "Excel", "立绘表.xlsx")
PORTRAITS_DIR = os.path.join(ROOT, "Assets", "Resources", "OutGameUI", "Portraits")

SHEET = "立绘"
HEADER = ["立绘ID", "资源路径", "备注"]
FIELDS = ["portraitId", "path", "note"]
WIDTH = [22, 40, 46]

HEAD_FILL = PatternFill("solid", fgColor="FFD9E2F3")
FIELD_FONT = Font(italic=True, color="FF808080", size=9)

# 种族中文名，纯粹为了让起始内容的备注好读；扫到别的素材就留空
RACE_NAMES = {"rabbit": "兔族", "goat": "羊族", "wolf": "狼族", "leopard": "豹族",
              "cheetah": "猎豹族", "ox": "牛族", "cat": "猫族", "yak": "牦牛族"}


def scan_portraits():
    """扫 Resources/OutGameUI/Portraits 下的立绘素材，生成 (立绘ID, 资源路径, 备注)。"""
    if not os.path.isdir(PORTRAITS_DIR):
        return []
    rows = []
    for name in sorted(os.listdir(PORTRAITS_DIR)):
        stem, ext = os.path.splitext(name)
        if ext.lower() not in (".png", ".jpg", ".jpeg"):
            continue
        note = "%s默认脸" % RACE_NAMES[stem] if stem in RACE_NAMES else ""
        rows.append(("%s_平静" % stem, "OutGameUI/Portraits/%s" % stem, note))
    return rows


def build_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    ws.append(HEADER)
    ws.append(FIELDS)
    for col in range(1, len(HEADER) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = HEAD_FILL
        ws.cell(row=2, column=col).font = FIELD_FONT
        ws.column_dimensions[get_column_letter(col)].width = WIDTH[col - 1]
    ws.freeze_panes = "A3"

    for row in rows:
        ws.append(list(row))
    return wb


def main():
    force = "--force" in sys.argv
    if os.path.exists(EXCEL_PATH) and not force:
        print("[SKIP] %s already exists. Pass --force to rebuild (this DISCARDS existing rows)."
              % os.path.abspath(EXCEL_PATH))
        return

    rows = scan_portraits()
    if not rows:
        print("[WARN] no portrait images found under %s" % PORTRAITS_DIR)

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    build_workbook(rows).save(EXCEL_PATH)
    print("[OK] wrote %s" % os.path.abspath(EXCEL_PATH))
    for pid, path, _ in rows:
        print("     %s -> %s" % (pid, path))


if __name__ == "__main__":
    main()
