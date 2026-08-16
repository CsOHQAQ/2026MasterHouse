# -*- coding: utf-8 -*-
"""一次性迁移：把访客阵容从 4 个占位种族（fox/crow/hedgehog/rabbit）换成美术交付的 8 个角色。

改的是**唯一数据源** Excel/*.xlsx（立绘表 / 访客种族表 / 访客日程表 / 对话表），
跑完再跑 export_config.bat（或单独跑 export_visitor.py + export_portrait.py + export_dialogue.py）
生成 Assets/Configs/*.csv，Unity 侧由 CsvPostprocessor 自动重建 SO。

阵容与素材的对应关系见 Tools/install_visitor_art.py。
对话内容是把原 fox 那一套按种族复制 8 份（文案先沿用，立绘ID 换成各自的脸），
保证每个角色点了都有话说；后续文案由策划直接在 Excel 里改。

对话组ID 分段：百位 = 种族序号（rabbit=0 → 10001/20001/30001…，goat=1 → 10101/20101/30101…）。
"""
import copy
import os

import openpyxl

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL = lambda name: os.path.join(ROOT, "Excel", name)

# (raceId, 显示名, 等搭话超时, 等交货超时, 闲逛上限, 备注)
CAST = [
    ("rabbit",  "兔族",   4200, 7200, 6000, "白兔·蓝斗篷，最能等"),
    ("goat",    "羊族",   3600, 6600, 5400, "白山羊·捧盆栽"),
    ("wolf",    "狼族",   1800, 3600, 2400, "白狼·白西装，急性子"),
    ("leopard", "豹族",   2400, 4800, 3000, "豹·棒球帽"),
    ("cheetah", "猎豹族", 2100, 4200, 2700, "猎豹·棒球服，坐不住"),
    ("ox",      "牛族",   4800, 8400, 6600, "褐牛·毛线帽，慢性子"),
    ("cat",     "猫族",   3000, 6000, 3600, "黑猫·警官"),
    ("yak",     "牦牛族", 3900, 7000, 5600, "白牦牛·蓝衬衫"),
]

# 日程：覆盖全部 8 个角色，需求沿用现有三条 NeedDef
SCHEDULE = [
    (1, 510, "rabbit",  "Need_修理电路"),
    (1, 550, "cat",     "Need_要蓝色椅子"),
    (1, 600, "goat",    "Need_新需求"),
    (1, 840, "wolf",    "Need_要蓝色椅子"),
    (2, 520, "ox",      "Need_修理电路"),
    (2, 570, "leopard", "Need_新需求"),
    (2, 780, "cheetah", "Need_要蓝色椅子"),
    (3, 540, "yak",     "Need_新需求"),
    (3, 660, "rabbit",  "Need_修理电路"),
    (3, 750, "cat",     "Need_要蓝色椅子"),
]

PORTRAIT_ID = lambda race: race + "_平静"
DATA_ROW = 3  # 第 1 行中文列名、第 2 行 Unity 字段名，数据从第 3 行起


def reset_rows(ws, keep_from=DATA_ROW):
    """清掉数据区，只留两行表头；样式从原第一行数据行取样，后面写新行时套上。"""
    style_row = [copy.copy(ws.cell(keep_from, c)._style) for c in range(1, ws.max_column + 1)]
    if ws.max_row >= keep_from:
        ws.delete_rows(keep_from, ws.max_row - keep_from + 1)
    return style_row


def write_row(ws, row, values, style_row):
    for i, value in enumerate(values, start=1):
        cell = ws.cell(row, i, value)
        if i - 1 < len(style_row):
            cell._style = copy.copy(style_row[i - 1])


def migrate_portrait():
    path = EXCEL("立绘表.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["立绘"]
    style = reset_rows(ws)
    for i, (race, display, *_rest, note) in enumerate(CAST):
        write_row(ws, DATA_ROW + i,
                  [PORTRAIT_ID(race), "OutGameUI/Portraits/" + race, display + "_" + note],
                  style)
    wb.save(path)
    print("立绘表：%d 行" % len(CAST))


def migrate_race():
    path = EXCEL("访客种族表.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["种族"]
    style = reset_rows(ws)
    for i, (race, display, talk, deliver, wander, _note) in enumerate(CAST):
        write_row(ws, DATA_ROW + i,
                  [race, display, talk, deliver, wander,
                   PORTRAIT_ID(race), "OutGameUI/Visitors/" + race],
                  style)
    wb.save(path)
    print("访客种族表：%d 行" % len(CAST))


def migrate_schedule():
    path = EXCEL("访客日程表.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["日程"]
    style = reset_rows(ws)
    for i, (day, minute, race, need) in enumerate(SCHEDULE):
        write_row(ws, DATA_ROW + i, [day, minute, race, need, None], style)
    wb.save(path)
    print("访客日程表：%d 行" % len(SCHEDULE))


def migrate_dialogue():
    """把 fox 那一套对话组按种族复制 8 份：组ID 加 种族序号×100，raceId / 立绘ID 换成各自的。"""
    path = EXCEL("对话表.xlsx")
    wb = openpyxl.load_workbook(path)

    groups = wb["对话组"]
    template = [[groups.cell(r, c).value for c in range(1, groups.max_column + 1)]
                for r in range(DATA_ROW, groups.max_row + 1)]
    style = reset_rows(groups)
    row = DATA_ROW
    for i, (race, *_rest) in enumerate(CAST):
        for values in template:
            values = list(values)
            values[0] = str(int(values[0]) + i * 100)  # 对话组ID
            values[1] = race                            # 种族
            write_row(groups, row, values, style)
            row += 1
    group_count = row - DATA_ROW

    lines = wb["对话内容"]
    template = [[lines.cell(r, c).value for c in range(1, lines.max_column + 1)]
                for r in range(DATA_ROW, lines.max_row + 1)]
    style = reset_rows(lines)
    row = DATA_ROW
    for i, (race, *_rest) in enumerate(CAST):
        for values in template:
            values = list(values)
            if values[0]:                                       # 对话组ID（只有每组首行有）
                values[0] = str(int(values[0]) + i * 100)
            if values[2]:                                       # 立绘ID（Branch/Action 行留空）
                values[2] = PORTRAIT_ID(race)
            write_row(lines, row, values, style)
            row += 1
    line_count = row - DATA_ROW

    # 参考页的下拉参考列：立绘ID 与 种族id 换成新阵容（纯给策划看的，不参与导表）
    ref = wb["参考"]
    for col, values in ((10, [PORTRAIT_ID(race) for race, *_ in CAST]),
                        (22, [race for race, *_ in CAST])):
        for r in range(2, ref.max_row + 1):
            ref.cell(r, col).value = None
        for i, value in enumerate(values):
            ref.cell(2 + i, col).value = value

    wb.save(path)
    print("对话表：对话组 %d 行、对话内容 %d 行" % (group_count, line_count))


if __name__ == "__main__":
    migrate_portrait()
    migrate_race()
    migrate_schedule()
    migrate_dialogue()
    print("Excel 已更新，请跑 Tools/导表/export_config.bat 生成 CSV。")
