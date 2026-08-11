# -*- coding: utf-8 -*-
"""将 Excel/家具房间表.xlsx（房间/网格/占用格/初始摆放 四张工作表）展平导出为 Unity 读取的 CSV。
数据唯一来源：Excel/家具房间表.xlsx；输出进 Assets/Configs/，Unity 资产管线检测到变化后自动重建 FurnitureRoomTable.asset。
CSV 为「记录类型」混排格式，房间行排在其明细行之前（Unity 导入器要求）。
"""
import os
import re
import sys

import openpyxl

# 表格第二行是「Unity 字段名」参考行（如 roomId / displayName），全 ASCII 标识符时跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Excel", "家具房间表.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Assets", "Configs", "家具房间表.csv")

# CSV 列（与 Unity 侧 FurnitureCsvImporter 表头一致，勿改文字）
HEADER = ["记录类型", "房间id", "显示名", "场景宽", "场景高", "背景图", "景深模糊图",
          "失焦模糊图", "初始货币", "网格id", "表面类型", "列数", "行数", "格宽", "格高",
          "X", "Y", "家具id", "宿主家具id", "列", "行", "翻转"]

# xlsx 各工作表自己的列（Excel 里列顺序可调，按表头名识别）
SHEET_COLUMNS = {
    "房间": ["房间id", "显示名", "场景宽", "场景高", "背景图", "景深模糊图", "失焦模糊图", "初始货币"],
    "网格": ["房间id", "网格id", "表面类型", "列数", "行数", "格宽", "格高", "X", "Y"],
    "占用格": ["房间id", "网格id", "列", "行"],
    "初始摆放": ["房间id", "家具id", "网格id", "宿主家具id", "列", "行", "翻转"],
}


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def csv_cell(text):
    if any(ch in text for ch in ',"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def read_sheet(workbook, name):
    if name not in workbook.sheetnames:
        print(f"[ERROR] sheet '{name}' missing in 家具房间表.xlsx")
        sys.exit(1)
    rows = list(workbook[name].iter_rows(values_only=True))
    if not rows:
        return []
    header = [cell_text(cell) for cell in rows[0]]
    missing = [col for col in SHEET_COLUMNS[name] if col not in header]
    if missing:
        print(f"[ERROR] sheet '{name}' missing columns: {missing}")
        sys.exit(1)
    index = {col: header.index(col) for col in SHEET_COLUMNS[name]}
    result = []
    for row_number, raw in enumerate(rows[1:]):
        record = {col: cell_text(raw[i]) if i < len(raw) else "" for col, i in index.items()}
        if not any(record.values()):
            continue
        if row_number == 0 and is_field_name_row(record.values()):
            continue  # 第二行 = Unity 字段名参考行
        result.append(record)
    return result


def export():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel not found: {os.path.abspath(EXCEL_PATH)}")
        sys.exit(1)

    workbook = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rooms = read_sheet(workbook, "房间")
    grids = read_sheet(workbook, "网格")
    blocked = read_sheet(workbook, "占用格")
    placements = read_sheet(workbook, "初始摆放")
    workbook.close()

    room_ids = [room["房间id"] for room in rooms]
    for kind, rows in (("网格", grids), ("占用格", blocked), ("初始摆放", placements)):
        for row in rows:
            if row["房间id"] not in room_ids:
                print(f"[WARN] {kind} 行引用了「房间」表里不存在的房间id: {row['房间id']}")

    lines = [",".join(HEADER)]

    def append(kind, record):
        merged = {"记录类型": kind, **record}
        lines.append(",".join(csv_cell(merged.get(col, "")) for col in HEADER))

    for room in rooms:
        room_id = room["房间id"]
        append("房间", room)
        for grid in grids:
            if grid["房间id"] == room_id:
                append("网格", grid)
        for cell in blocked:
            if cell["房间id"] == room_id:
                append("占用格", cell)
        for place in placements:
            if place["房间id"] == room_id:
                append("初始摆放", place)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8-sig", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    print(f"[OK] Exported {len(rooms)} rooms, {len(lines) - 1} records -> {os.path.abspath(OUTPUT_PATH)}")


if __name__ == "__main__":
    export()
