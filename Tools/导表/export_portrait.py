# -*- coding: utf-8 -*-
"""立绘表导出：Excel/立绘表.xlsx → Assets/Configs/立绘表.csv

策划编辑 Excel/立绘表.xlsx（工作表「立绘」）
→ 双击 Tools/导表/export_config.bat
→ CSV 写进 Assets/Configs/
→ Unity 资产管线检测到变化，PortraitCsvImporter 整表重建
   Assets/Resources/OutGameUI/PortraitTable.asset，**并顺带重导一次对话表**
   （对话表要拿立绘表校验「立绘ID 存不存在」，两者必须同批更新）。

这张表就两件事：给每张立绘一个 ID，记下它在 Resources 里的路径。
表情枚举已于 2026-08-14 退役——差分数量与命名完全交给美术，加一张图 = 加一行，不碰代码。

**没有反向导出**：SO 是产物不是源。要看现有内容就打开 xlsx。

本脚本的校验只到「格式层」（ID 空不空、重不重、路径填没填）；
「路径指向的贴图在不在」要读 Unity 工程，由 PortraitCsvImporter 负责（那边只给警告，
不拦导表——美术流程里「先占 ID 后补图」是常态）。

**导出时会追加一列「行号」**：值是 Excel 里的真实行号，Unity 侧报错靠它指回策划能改的地方。
"""
import csv
import os
import re
import sys

import openpyxl

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL_PATH = os.path.join(ROOT, "Excel", "立绘表.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "Assets", "Configs")

SHEET = "立绘"
HEADER = ["立绘ID", "资源路径", "备注"]

# 第二行是「Unity 字段名」参考行（全 ASCII 标识符），与其它表一致地跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")

errors = []


def fail(row, message):
    errors.append("[%s] row %d: %s" % (SHEET, row, message))


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))  # Excel 把整数存成 1.0，导出时还原
    return str(value).strip()


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)


def read_sheet(wb):
    """返回 [(excel_row, {列名: 值}), ...]；表头缺列直接退出（那是模板问题，不是内容问题）。"""
    if SHEET not in wb.sheetnames:
        print("[ERROR] sheet '%s' not found in %s" % (SHEET, os.path.basename(EXCEL_PATH)))
        sys.exit(1)

    rows = list(wb[SHEET].iter_rows(values_only=True))
    if not rows:
        print("[ERROR] sheet '%s' is empty" % SHEET)
        sys.exit(1)

    actual = [cell_text(v) for v in rows[0]]
    missing = [c for c in HEADER if c not in actual]
    if missing:
        print("[ERROR] sheet '%s' missing columns: %s" % (SHEET, missing))
        sys.exit(1)
    index = {name: actual.index(name) for name in HEADER}

    result = []
    seen_data = False
    for offset, raw in enumerate(rows[1:]):
        excel_row = offset + 2
        values = [cell_text(v) for v in raw]
        values += [""] * (len(actual) - len(values))
        if not any(values):
            continue
        if not seen_data and is_field_name_row(values):
            continue  # 第二行英文字段名参考行
        seen_data = True
        result.append((excel_row, {name: values[i] if i < len(values) else "" for name, i in index.items()}))
    return result


def validate(rows):
    seen = {}
    for row, data in rows:
        pid = data["立绘ID"]
        path = data["资源路径"]
        if not pid:
            fail(row, "portrait id is empty (path: %s)" % (path or "<empty>"))
            continue
        # 逗号是 CSV 的分隔符，虽然导出会加引号保护，但 ID 里带逗号纯属自找麻烦
        if "," in pid:
            fail(row, "portrait id must not contain a comma: %s" % pid)
        if pid in seen:
            fail(row, "duplicate portrait id '%s', already used at row %d" % (pid, seen[pid]))
        else:
            seen[pid] = row
        if not path:
            fail(row, "portrait '%s' has no resource path" % pid)


def write_csv(rows):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, "立绘表.csv")
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADER + ["行号"])
        for row, data in rows:
            writer.writerow([data[c] for c in HEADER] + [row])
    print("[OK] 立绘表.csv: %d rows" % len(rows))


def main():
    if not os.path.exists(EXCEL_PATH):
        print("[ERROR] Excel not found: %s" % os.path.abspath(EXCEL_PATH))
        print("        Run Tools/导表/make_portrait_template.py to create a blank template.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    rows = read_sheet(wb)
    wb.close()

    validate(rows)
    if errors:
        print("[ERROR] portrait table has %d problem(s); nothing was written:" % len(errors))
        for message in errors:
            print("        " + message)
        sys.exit(1)

    write_csv(rows)


if __name__ == "__main__":
    main()
