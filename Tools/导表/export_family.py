# -*- coding: utf-8 -*-
"""将 Excel/家具族表.xlsx（工作表「家具族」）导出为 Unity 读取的 CSV。
族级共有属性（分类/表面/占格/装饰分/音效/桌面格）的唯一来源：改一处、整族生效。
Unity 侧导入时按家具表每行的「族id」把这些列**展开填进** FurnitureEntry，所以本表必须先于家具表导入。
"""
import os
import re
import sys

import openpyxl

# 表格第二行是「Unity 字段名」参考行（如 familyId / displayName），全 ASCII 标识符时跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Excel", "家具族表.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Assets", "Configs", "家具族表.csv")

# CSV 列（与 Unity 侧 FurnitureCsvImporter.FamilyHeader 一致，勿改文字）
HEADER = ["族id", "族显示名", "分类", "描述", "表面类型", "可叠放", "占格列", "占格行", "装饰分",
          "拿起音效", "放下音效", "桌面格启用", "桌面格列数", "桌面格宽", "桌面格高", "桌面格偏移X", "桌面高度"]
SHEET = "家具族"


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))  # Excel 把整数存成 1.0，导出时还原
    return str(value).strip()


def csv_cell(text):
    if any(ch in text for ch in ',"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def export():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel not found: {os.path.abspath(EXCEL_PATH)}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"[ERROR] sheet '{SHEET}' missing in 家具族表.xlsx")
        sys.exit(1)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        print("[ERROR] 家具族表.xlsx is empty or has no data rows.")
        sys.exit(1)

    header = [cell_text(cell) for cell in rows[0]]
    missing = [col for col in HEADER if col not in header]
    if missing:
        print(f"[ERROR] 家具族表.xlsx missing columns: {missing}")
        sys.exit(1)
    index = {col: header.index(col) for col in HEADER}

    lines = [",".join(HEADER)]
    seen = set()
    count = 0
    for row_number, raw in enumerate(rows[1:]):
        record = {col: cell_text(raw[i]) if i < len(raw) else "" for col, i in index.items()}
        if not any(record.values()):
            continue
        if row_number == 0 and is_field_name_row(record.values()):
            continue  # 第二行 = Unity 字段名参考行
        if not record["族id"]:
            print(f"[WARN] 家具族表 存在缺 族id 的行，已跳过: {record}")
            continue
        if record["族id"] in seen:
            print(f"[WARN] 家具族表 族id 重复，后一行已跳过: {record['族id']}")
            continue
        seen.add(record["族id"])
        lines.append(",".join(csv_cell(record[col]) for col in HEADER))
        count += 1

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8-sig", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    print(f"[OK] Exported {count} furniture family rows -> {os.path.abspath(OUTPUT_PATH)}")


if __name__ == "__main__":
    export()
