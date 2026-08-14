# -*- coding: utf-8 -*-
"""将访客三张 Excel 导出为 Unity 读取的 CSV。
数据唯一来源：Excel/访客种族表.xlsx、访客日程表.xlsx、访客调参表.xlsx；
输出进 Assets/Configs/，Unity 资产管线检测到变化后由 VisitorCsvImporter 自动重建对应 SO
（VisitorRaces/Race_*.asset / VisitorScheduleTable / VisitorTuningConfig）。
"""
import os
import re
import sys

import openpyxl

# 表格第二行是「Unity 字段名」参考行（如 raceId / displayName），全 ASCII 标识符时跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)


ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL_DIR = os.path.join(ROOT, "Excel")
OUTPUT_DIR = os.path.join(ROOT, "Assets", "Configs")

# 每项：(xlsx 文件, 工作表, CSV 输出, 必须包含的列)——列名与 Unity 侧 VisitorCsvImporter 一致，勿改文字
JOBS = [
    # 需求权重 / 需求数上下限三列已随 tag 需求体系退役（访客需求重做说明 §9.1）：
    # 谁带什么需求来改由访客日程表的「需求」列逐条配死。
    # 「对话池」列已随 2026-08-14 对话资源重构退役：对话内容改由 Excel/对话表.xlsx 一张整表承载，
    # 按 raceId 关联，种族资产上不再挂对话池引用。
    # 「立绘差分」列已随 2026-08-14 立绘 ID 化退役，换成单列「默认立绘ID」（写 Excel/立绘表.xlsx 的主键）：
    # 原来那种 `平静=路径/高兴=路径` 的双层分隔串正是 §16.6 禁止的无类型数据，差分整体搬去立绘表。
    # 「跨天留宿概率%」列同批删除：消费方（闲逛访客的跨天 roll）早已移除，见 VisitorManager.EndDay。
    ("访客种族表.xlsx", "种族", "访客种族表.csv",
     ["种族id", "显示名", "等搭话超时tick", "等交货超时tick", "闲逛上限tick",
      "默认立绘ID", "序列帧"]),
    ("访客日程表.xlsx", "日程", "访客日程表.csv",
     ["天", "出现时刻(分钟)", "种族id", "需求", "具名覆写"]),
    ("访客调参表.xlsx", "调参", "访客调参表.csv",
     ["参数", "值"]),
    ("访客调参表.xlsx", "氛围访客", "访客氛围表.csv",
     ["id", "显示名", "序列帧"]),
]


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))  # Excel 把整数存成 1.0，导出时还原
    return str(value).strip()


def csv_cell(text):
    if any(ch in text for ch in ',"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def export(xlsx, sheet, csv_name, required):
    excel_path = os.path.join(EXCEL_DIR, xlsx)
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel not found: {os.path.abspath(excel_path)}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"[ERROR] sheet '{sheet}' missing in {xlsx}")
        sys.exit(1)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        print(f"[ERROR] {xlsx}[{sheet}] is empty or has no data rows.")
        sys.exit(1)

    header = [cell_text(cell) for cell in rows[0]]
    missing = [col for col in required if col not in header]
    if missing:
        print(f"[ERROR] {xlsx}[{sheet}] missing columns: {missing}")
        sys.exit(1)

    lines = [",".join(csv_cell(h) for h in header)]
    count = 0
    for row in rows[1:]:
        values = [cell_text(cell) for cell in row[:len(header)]]
        values += [""] * (len(header) - len(values))
        if not any(values):
            continue
        if count == 0 and is_field_name_row(values):
            continue  # 第二行英文字段名参考行
        lines.append(",".join(csv_cell(v) for v in values))
        count += 1

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, csv_name)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    print(f"[OK] {csv_name}: {count} rows")


def main():
    for xlsx, sheet, csv_name, required in JOBS:
        export(xlsx, sheet, csv_name, required)


if __name__ == "__main__":
    main()
