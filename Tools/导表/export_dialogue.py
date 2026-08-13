# -*- coding: utf-8 -*-
"""对话表导出：Excel/对话表.xlsx → Assets/Configs/对话组表.csv + 对话内容表.csv

策划编辑 Excel/对话表.xlsx（两个 Sheet：对话组 / 对话内容）
→ 双击 Tools/导表/export_config.bat
→ CSV 写进 Assets/Configs/
→ Unity 资产管线检测到变化，DialogueCsvImporter 整表重建 Assets/Resources/OutGameUI/DialogueTable.asset。

**没有反向导出**：SO 是产物不是源。要看现有内容就打开 xlsx。

本脚本做的校验只到「格式层」（列在不在、ID 是不是数字、类型合不合法、步骤号有没有重复），
够不够得着业务（函数名存不存在、需求有没有配对话）由 Unity 侧的导入器与校验器负责——
那些要读工程资产，Python 这边看不见。

**导出时会给每张表追加一列「行号」**：值是 Excel 里的真实行号，
Unity 侧报错时靠它把问题指回策划能改的地方。
"""
import csv
import os
import re
import sys

import openpyxl

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL_PATH = os.path.join(ROOT, "Excel", "对话表.xlsx")
OUTPUT_DIR = os.path.join(ROOT, "Assets", "Configs")

GROUP_SHEET = "对话组"
CONTENT_SHEET = "对话内容"

GROUP_HEADER = ["对话组ID", "种族", "需求ID", "所属对话池", "进入条件", "备注"]
CONTENT_HEADER = ["对话组ID", "说话人", "表情", "步骤", "选项", "句序", "类型", "文本", "条件"]

CATEGORIES = [
    "firstMeeting", "waitingReception", "needTalk",
    "feedbackDisappointed", "feedbackPlain", "feedbackFine", "feedbackPerfect",
    "smallTalk",
]
KINDS = ["Line", "Action", "Branch"]
SPEAKERS = ["visitor", "player", "narration"]
EMOTIONS = ["calm", "happy", "confused", "sad", "surprised"]

# 第二行是「Unity 字段名」参考行（全 ASCII 标识符），与其它表一致地跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")

errors = []


def fail(sheet, row, message):
    errors.append("[%s] row %d: %s" % (sheet, row, message))


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))  # Excel 把整数存成 1.0，导出时还原
    return str(value).strip()


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)


def read_sheet(wb, sheet, header):
    """返回 [(excel_row, {列名: 值}), ...]；表头缺列直接退出（那是模板问题，不是内容问题）。"""
    if sheet not in wb.sheetnames:
        print("[ERROR] sheet '%s' not found in %s" % (sheet, os.path.basename(EXCEL_PATH)))
        sys.exit(1)

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("[ERROR] sheet '%s' is empty" % sheet)
        sys.exit(1)

    actual = [cell_text(v) for v in rows[0]]
    missing = [c for c in header if c not in actual]
    if missing:
        print("[ERROR] sheet '%s' missing columns: %s" % (sheet, missing))
        sys.exit(1)
    index = {name: actual.index(name) for name in header}

    result = []
    seen_data = False
    for offset, raw in enumerate(rows[1:]):
        excel_row = offset + 2
        values = [cell_text(v) for v in raw]
        values += [""] * (len(actual) - len(values))
        if not any(values):
            continue
        if not seen_data and is_field_name_row(values):
            continue  # 第二行英文字段名参考行
        seen_data = True
        result.append((excel_row, {name: values[i] if i < len(values) else "" for name, i in index.items()}))
    return result


def check_int(sheet, row, value, label, required=True):
    if value == "":
        if required:
            fail(sheet, row, "%s is empty" % label)
        return None
    try:
        return int(value)
    except ValueError:
        fail(sheet, row, "%s is not an integer: %s" % (label, value))
        return None


def validate_groups(rows):
    for row, data in rows:
        check_int(GROUP_SHEET, row, data["对话组ID"], "group id")
        if data["种族"] == "":
            fail(GROUP_SHEET, row, "race is empty (use raceId, `/` for multi, or 通用)")
        category = data["所属对话池"]
        if category not in CATEGORIES:
            fail(GROUP_SHEET, row, "unknown category: %s (expected one of %s)" % (category, "/".join(CATEGORIES)))
        elif category == "needTalk" and data["需求ID"] == "":
            fail(GROUP_SHEET, row, "needTalk requires a 需求ID")
        elif data["需求ID"] != "" and not category.startswith("feedback") and category != "needTalk":
            fail(GROUP_SHEET, row, "category %s must leave 需求ID empty" % category)


def fill_down_group_id(rows):
    """第二页的「对话组ID」只在每组首行写一次，后续行留空 = 沿用上一行（策划的写法）。

    在这里**就地补全**，让 CSV 与 Unity 侧永远拿到具体 ID——
    「留空继承」这种依赖行序的规则只应该活在 Excel 那一层，不该漏进下游。
    注意：正因为它依赖行序，**第二页不要在 Excel 里排序**；三列数字序号已经保证了
    结构不依赖行序，排序只会打乱这一列的继承。
    """
    current = ""
    for row, data in rows:
        value = data["对话组ID"]
        if value:
            current = value
        elif current:
            data["对话组ID"] = current
        else:
            fail(CONTENT_SHEET, row, "group id is empty and there is no previous row to inherit from")


def validate_content(rows):
    # (group, step, option, sub) 唯一；option/sub 空视为 -1
    seen = {}
    for row, data in rows:
        gid = check_int(CONTENT_SHEET, row, data["对话组ID"], "group id")
        step = check_int(CONTENT_SHEET, row, data["步骤"], "step")
        option = check_int(CONTENT_SHEET, row, data["选项"], "option", required=False)
        sub = check_int(CONTENT_SHEET, row, data["句序"], "sub", required=False)
        kind = data["类型"]
        if kind not in KINDS:
            fail(CONTENT_SHEET, row, "unknown kind: %s (expected one of %s)" % (kind, "/".join(KINDS)))
        if data["说话人"] and data["说话人"] not in SPEAKERS:
            fail(CONTENT_SHEET, row, "unknown speaker: %s" % data["说话人"])
        if data["表情"] and data["表情"] not in EMOTIONS:
            fail(CONTENT_SHEET, row, "unknown emotion: %s" % data["表情"])
        if kind == "Branch" and option is None:
            fail(CONTENT_SHEET, row, "Branch row must fill the 选项 column (option index, starting at 1)")
        if sub is not None and option is None:
            fail(CONTENT_SHEET, row, "句序 is filled but 选项 is empty")
        if kind == "Branch" and sub is not None:
            fail(CONTENT_SHEET, row, "nested branches are not supported (Branch row must leave 句序 empty)")
        if kind == "Action" and data["文本"] == "":
            fail(CONTENT_SHEET, row, "Action row needs a call in the 文本 column, e.g. Accept")

        if gid is None or step is None:
            continue
        key = (gid, step, option if option is not None else -1, sub if sub is not None else -1)
        if key in seen:
            fail(CONTENT_SHEET, row, "duplicate position, same as row %d" % seen[key])
        else:
            seen[key] = row


def write_csv(name, header, rows):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, name)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header + ["行号"])
        for row, data in rows:
            writer.writerow([data[c] for c in header] + [row])
    print("[OK] %s: %d rows" % (name, len(rows)))


def main():
    if not os.path.exists(EXCEL_PATH):
        print("[ERROR] Excel not found: %s" % os.path.abspath(EXCEL_PATH))
        print("        Run Tools/导表/make_dialogue_template.py to create a blank template.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    groups = read_sheet(wb, GROUP_SHEET, GROUP_HEADER)
    content = read_sheet(wb, CONTENT_SHEET, CONTENT_HEADER)
    wb.close()

    validate_groups(groups)
    fill_down_group_id(content)
    validate_content(content)
    if errors:
        print("[ERROR] dialogue table has %d problem(s); nothing was written:" % len(errors))
        for message in errors:
            print("        " + message)
        sys.exit(1)

    write_csv("对话组表.csv", GROUP_HEADER, groups)
    write_csv("对话内容表.csv", CONTENT_HEADER, content)


if __name__ == "__main__":
    main()
