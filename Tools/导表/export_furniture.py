# -*- coding: utf-8 -*-
"""将 Excel/家具表.xlsx（工作表「家具」）导出为 Unity 读取的 CSV。
数据唯一来源：Excel/家具表.xlsx；输出进 Assets/Configs/，Unity 资产管线检测到变化后自动重建 FurnitureTable.asset。
"""
import os
import re
import sys

import openpyxl

# 表格第二行是「Unity 字段名」参考行（如 id / displayName），全 ASCII 标识符时跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Excel", "家具表.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Assets", "Configs", "家具表.csv")

# CSV 列（与 Unity 侧 FurnitureCsvImporter 表头一致，勿改文字）
HEADER = ["id", "显示名", "分类", "描述", "表面类型", "占格列", "占格行", "显示宽", "显示高", "价格",
          "解禁声望", "装饰分", "精灵图", "桌面格启用", "桌面格列数", "桌面格宽",
          "桌面格高", "桌面格偏移X", "桌面高度"]
SHEET = "家具"


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))  # Excel 把整数存成 1.0，导出时还原
    return str(value).strip()


def csv_cell(text):
    if any(ch in text for ch in ',"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def export():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel not found: {os.path.abspath(EXCEL_PATH)}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"[ERROR] sheet '{SHEET}' missing in 家具表.xlsx")
        sys.exit(1)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        print("[ERROR] 家具表.xlsx is empty or has no data rows.")
        sys.exit(1)

    header = [cell_text(cell) for cell in rows[0]]
    missing = [col for col in HEADER if col not in header]
    if missing:
        print(f"[ERROR] 家具表.xlsx missing columns: {missing}")
        sys.exit(1)
    index = {col: header.index(col) for col in HEADER}

    lines = [",".join(HEADER)]
    count = 0
    for row_number, raw in enumerate(rows[1:]):
        record = {col: cell_text(raw[i]) if i < len(raw) else "" for col, i in index.items()}
        if not any(record.values()):
            continue
        if row_number == 0 and is_field_name_row(record.values()):
            continue  # 第二行 = Unity 字段名参考行
        if not record["id"]:
            print(f"[WARN] 家具表 存在缺 id 的行，已跳过: {record}")
            continue
        lines.append(",".join(csv_cell(record[col]) for col in HEADER))
        count += 1

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8-sig", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    print(f"[OK] Exported {count} furniture rows -> {os.path.abspath(OUTPUT_PATH)}")


if __name__ == "__main__":
    export()
