# -*- coding: utf-8 -*-
"""将 Excel/商店表.xlsx（工作表「商店」）导出为 Unity 读取的 CSV。
商店表是家具的售卖配置（价格 / 解禁声望），按 id 关联家具表；
导入时由 FurnitureCsvImporter 合回 FurnitureTable.asset 的对应条目。
"""
import os
import re
import sys

import openpyxl

# 表格第二行是「Unity 字段名」参考行（如 id / price），全 ASCII 标识符时跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")


def is_field_name_row(values):
    filled = [v for v in values if v]
    return bool(filled) and all(FIELD_NAME.match(v) for v in filled)


EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Excel", "商店表.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "Assets", "Configs", "商店表.csv")

# CSV 列（与 Unity 侧 FurnitureCsvImporter 表头一致，勿改文字）
HEADER = ["id", "显示名", "价格", "解禁声望"]
SHEET = "商店"


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def csv_cell(text):
    if any(ch in text for ch in ',"\n'):
        return '"' + text.replace('"', '""') + '"'
    return text


def export():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel not found: {os.path.abspath(EXCEL_PATH)}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"[ERROR] sheet '{SHEET}' missing in 商店表.xlsx")
        sys.exit(1)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        print("[ERROR] 商店表.xlsx is empty or has no data rows.")
        sys.exit(1)

    header = [cell_text(cell) for cell in rows[0]]
    missing = [col for col in HEADER if col not in header]
    if missing:
        print(f"[ERROR] 商店表.xlsx missing columns: {missing}")
        sys.exit(1)

    lines = [",".join(csv_cell(h) for h in header)]
    count = 0
    for row in rows[1:]:
        values = [cell_text(cell) for cell in row[:len(header)]]
        values += [""] * (len(header) - len(values))
        if not any(values):
            continue
        if count == 0 and is_field_name_row(values):
            continue  # 第二行英文字段名参考行
        lines.append(",".join(csv_cell(v) for v in values))
        count += 1

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8-sig", newline="") as f:
        f.write("\r\n".join(lines) + "\r\n")
    print(f"[OK] 商店表.csv: {count} rows")


if __name__ == "__main__":
    export()
