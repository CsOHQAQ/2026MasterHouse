# -*- coding: utf-8 -*-
"""生成 Excel/对话表.xlsx 模板（两页 + 一页参考），并铺一套能直接跑起来的起始内容。

用法：
    python Tools/导表/make_dialogue_template.py          # 文件已存在时拒绝覆盖
    python Tools/导表/make_dialogue_template.py --force  # 强制重建（会丢掉现有台词！）

生成之后正常流程就是「改 xlsx → 双击 export_config.bat → 切回 Unity」，本脚本不再需要。

模板里做了三件让策划省事的事：
  1. 关键列全部配数据校验下拉（对话池 / 说话人 / 立绘ID / 类型 / 需求ID / 种族），
     从下拉里选，拼错这件事在填表阶段就不会发生
  2. 「参考」页列出全部可选值与可用函数，下拉直接引用它的区域——
     加了新种族/新需求/新立绘，重跑本脚本或手工改「参考」页即可
  3. 起始内容是**完整可跑**的一套：覆盖校验器要求的全部分类，以及工程里每一条需求的【需求对话】

⚠ 「参考」页里的条件/事件清单是 Assets/Scripts/Dialogue/DialogueFuncs.cs 的镜像。
   那两张字典才是唯一真相源——加了函数记得同步这里；忘了也不会出事故，
   只是下拉里选不到，而写错的名字导表时会被 Unity 侧当场报错。

【2026-08-14 立绘 ID 化的两处结构变化】
  · 「表情」列 → 「立绘ID」列：可选值不再是固定五项枚举，而是 Excel/立绘表.xlsx 里的行。
    **留空 = 沿用上一句**，所以起始内容只在每组首句写一次 ID。
  · 种族列不再有「通用」：一个对话组只属于一个种族，否则多个种族共用同一份台词
    就会共用同一个立绘ID（串脸）。于是起始内容按种族各生成一套，组数 = 原来的组数 × 种族数。
"""
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
EXCEL_PATH = os.path.join(ROOT, "Excel", "对话表.xlsx")
PORTRAIT_EXCEL_PATH = os.path.join(ROOT, "Excel", "立绘表.xlsx")
NEED_DIR = os.path.join(ROOT, "Assets", "GameData", "Needs")
RACE_DIR = os.path.join(ROOT, "Assets", "Resources", "OutGameUI", "VisitorRaces")

GROUP_SHEET = "对话组"
CONTENT_SHEET = "对话内容"
REF_SHEET = "参考"

GROUP_HEADER = ["对话组ID", "种族", "需求ID", "所属对话池", "进入条件", "备注"]
GROUP_FIELDS = ["groupId", "raceId", "needId", "category", "conditions", "note"]
GROUP_WIDTH = [11, 16, 20, 22, 30, 46]

CONTENT_HEADER = ["对话组ID", "说话人", "立绘ID", "步骤", "选项", "句序", "类型", "文本", "条件"]
CONTENT_FIELDS = ["groupId", "speaker", "portraitId", "step", "option", "sub", "kind", "text", "conditions"]
CONTENT_WIDTH = [11, 10, 18, 7, 7, 7, 9, 58, 26]

CATEGORIES = [
    ("firstMeeting", "初次见面：首次点击前台队首访客。组内要有「接待 / 拒绝」分支"),
    ("waitingReception", "等待接待：同一位前台访客的二次点击。**同样要带接待/拒绝分支**"),
    ("needTalk", "需求对话：入住并示意后点击他。说需求 + 交付/推迟/放弃分支。**需求ID 必填**"),
    ("feedbackDisappointed", "需求反馈·失望：服务超时，需求没办到"),
    ("feedbackPlain", "需求反馈·一般：小游戏低分。条件类走不到这一档"),
    ("feedbackFine", "需求反馈·还行：小游戏中间分。条件类走不到这一档"),
    ("feedbackPerfect", "需求反馈·完美：条件类交付成功 / 小游戏满分"),
    ("smallTalk", "闲聊：停留期冒泡，走场景气泡。**一组只显示第一句**，多句请拆成多组"),
    ("farewell", "告别：停留到点后转「待告别」，玩家点他才播。**组末尾必须配一行 Action|Leave**，"
                 "否则客人道完别也不会走，会一直占着那间客房"),
]
KINDS = [
    ("Line", "说一句话。填说话人 / 立绘ID / 文本"),
    ("Action", "执行事件。「文本」列写调用串（如 Accept），说话人与立绘ID 留空"),
    ("Branch", "一个选项。填「选项」列（从 1 起），「文本」是选项文字，「条件」是它的可选条件"),
]
SPEAKERS = [("visitor", "访客：显示立绘与名字条"), ("player", "玩家：无立绘、另一种框"),
            ("narration", "旁白：居中无框")]

# ── DialogueFuncs.cs 的镜像（见文件头注释）──
CONDITIONS = [
    ("HasEmptyRoom", "还有空客房（只看房，不看队列）"),
    ("CanAcceptGuest", "现在能接待新客人（有空房 且 没有别人正在等分房）"),
    ("RoomHasNeedFurniture", "所住房间里摆着他要的家具之一（条件类需求的验收判据）"),
    ("DayAtLeast(N)", "第 N 天或更晚"),
    ("CurrencyAtLeast(N)", "货币不少于 N"),
    ("ReputationAtLeast(N)", "声望不少于 N"),
    ("SatisfactionAtLeast(档)", "本次满意度不低于 disappointed/plain/fine/perfect"),
    ("VisitorStateIs(状态)", "访客处于 FrontDesk/AwaitingRoom/Serving/Wandering/AwaitingFarewell"),
]
ACTIONS = [
    ("Accept", "接待（转「等待分配房间」，此时不说需求）"),
    ("Reject", "拒绝（扣声望并离场）"),
    ("Leave", "送别离场（客人当场离开）。**只在【告别】组里有效，且必须是这条路径的最后一个事件**"),
    ("CompleteNeed(档)", "完成需求结算，留空 = perfect。**奖励类：必须是这条路径的最后一个事件**"),
    ("StartMinigame", "开始小游戏（小游戏类需求的开局口）"),
    ("AddCurrency(N)", "增减货币。**奖励类**"),
    ("AddReputation(N)", "增减声望。**奖励类**"),
    ("GrantFurniture(家具id,数量)", "赠送家具进库存，数量可省略 = 1。**奖励类**"),
    ("Log(文本)", "往 Console 打一条日志，自查分支走向用"),
]

HEAD_FILL = PatternFill("solid", fgColor="FFD9E2F3")
REF_FILL = PatternFill("solid", fgColor="FFEDEDED")
FIELD_FONT = Font(italic=True, color="FF808080", size=9)

# 各表第 2 行是「Unity 字段名」参考行（全 ASCII 标识符），读别的表时要跳过
FIELD_NAME = re.compile(r"^[A-Za-z][A-Za-z0-9_.]*$")


def scan_needs():
    """工程里的 NeedDef 资产名（第一页「需求ID」列填的就是它）。"""
    if not os.path.isdir(NEED_DIR):
        return []
    names = [os.path.splitext(f)[0] for f in os.listdir(NEED_DIR) if f.endswith(".asset")]
    names.sort()
    return names


def scan_races():
    """工程里的 raceId（从 VisitorRaceDef 资产的 YAML 里读）。"""
    if not os.path.isdir(RACE_DIR):
        return []
    pattern = re.compile(r"^\s*raceId:\s*(\S+)\s*$", re.M)
    ids = []
    for name in sorted(os.listdir(RACE_DIR)):
        if not name.endswith(".asset"):
            continue
        with open(os.path.join(RACE_DIR, name), encoding="utf-8", errors="ignore") as f:
            match = pattern.search(f.read())
        if match and match.group(1) not in ids:
            ids.append(match.group(1))
    return ids


def scan_portraits():
    """Excel/立绘表.xlsx 里的立绘ID（第二页「立绘ID」列的下拉源）。"""
    if not os.path.exists(PORTRAIT_EXCEL_PATH):
        return []
    wb = openpyxl.load_workbook(PORTRAIT_EXCEL_PATH, read_only=True, data_only=True)
    if "立绘" not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb["立绘"].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = ["" if v is None else str(v).strip() for v in rows[0]]
    if "立绘ID" not in header:
        return []
    index = header.index("立绘ID")
    ids = []
    seen_data = False
    for raw in rows[1:]:
        values = ["" if v is None else str(v).strip() for v in raw]
        if not any(values):
            continue
        # 第 2 行是「Unity 字段名」参考行（portraitId / path / note），不是数据
        filled = [v for v in values if v]
        if not seen_data and filled and all(FIELD_NAME.match(v) for v in filled):
            continue
        seen_data = True
        if index < len(values) and values[index] and values[index] not in ids:
            ids.append(values[index])
    return ids


def default_portrait_of(race_id, portraits):
    """这个种族的起始立绘ID：优先 `<raceId>_平静`，否则拿第一个以 raceId 开头的，再否则留空。"""
    preferred = "%s_平静" % race_id
    if preferred in portraits:
        return preferred
    for pid in portraits:
        if pid.startswith(race_id):
            return pid
    return ""


# ── 起始内容 ────────────────────────────────────────────────────────────────
# 一组 = (对话组ID, [行...])；行 = (说话人, 立绘ID, 步骤, 选项, 句序, 类型, 文本, 条件)
#
# 立绘ID 只在**每组首句**写一次：留空 = 沿用上一句，所以后面的行不必重复填。
# 现阶段每个种族只有一张立绘，等美术出了差分，在需要换脸的那一句补 ID 即可。

def first_meeting_rows(face):
    # ⚠ 第三个选项「（先想一下）」不是凑数的：接待与拒绝**都会让访客离开前台**
    # （Accept → 待分房、Reject → 离场），而 ESC 中断不算「正常播完」、不置 MetPlayer。
    # 少了这个「什么都不做但正常收尾」的出口，就不存在「前台 + 已打过招呼」的访客，
    # 【等待接待】这一整类永远抽不到——写了台词没人听得见。
    return [
        ("visitor", face, 1, "", "", "Line", "你好，打扰了……请问今晚还有空房吗？", ""),
        ("", "", 2, 1, "", "Branch", "有的，请进", "CanAcceptGuest"),
        ("", "", 2, 1, 1, "Action", "Accept", ""),
        ("visitor", "", 2, 1, 2, "Line", "太好了，谢谢你！", ""),
        ("", "", 2, 2, "", "Branch", "抱歉，今天恐怕招待不了你", ""),
        ("", "", 2, 2, 1, "Action", "Reject", ""),
        ("visitor", "", 2, 2, 2, "Line", "这样啊……那我改天再来。", ""),
        ("", "", 2, 3, "", "Branch", "（先想一下）", ""),
        ("visitor", "", 2, 3, 1, "Line", "好的，我在这儿等着。", ""),
    ]


def waiting_rows(face):
    return [
        ("visitor", face, 1, "", "", "Line", "我还能再等一会儿，你先忙。", ""),
        ("", "", 2, 1, "", "Branch", "久等了，请进", "CanAcceptGuest"),
        ("", "", 2, 1, 1, "Action", "Accept", ""),
        ("visitor", "", 2, 1, 2, "Line", "谢谢你！", ""),
        ("", "", 2, 2, "", "Branch", "抱歉，今天恐怕招待不了你", ""),
        ("", "", 2, 2, 1, "Action", "Reject", ""),
        ("visitor", "", 2, 2, 2, "Line", "这样啊……那我改天再来。", ""),
        ("", "", 2, 3, "", "Branch", "（先不打扰他）", ""),
    ]


def need_rows(face, minigame):
    if minigame:
        return [
            ("visitor", face, 1, "", "", "Line", "{需求}", ""),
            ("", "", 2, 1, "", "Branch", "我来看看", ""),
            ("", "", 2, 1, 1, "Action", "StartMinigame", ""),
            ("", "", 2, 2, "", "Branch", "等我准备一下", ""),
            ("visitor", "", 2, 2, 1, "Line", "好，我等着。", ""),
        ]
    return [
        ("visitor", face, 1, "", "", "Line", "{需求}", ""),
        ("", "", 2, 1, "", "Branch", "我把你要的搬来了", "RoomHasNeedFurniture"),
        ("visitor", "", 2, 1, 1, "Line", "就是这个！太谢谢你了。", ""),
        ("", "", 2, 1, 2, "Action", "CompleteNeed(perfect)", ""),
        ("", "", 2, 2, "", "Branch", "我这就去弄", ""),
        ("visitor", "", 2, 2, 1, "Line", "好，我不急，你慢慢来。", ""),
        ("", "", 2, 3, "", "Branch", "抱歉，这个我办不到", ""),
        ("visitor", "", 2, 3, 1, "Line", "唉……那也没办法。", ""),
        ("", "", 2, 3, 2, "Action", "Reject", ""),
    ]


def one_liner(face, text):
    return [("visitor", face, 1, "", "", "Line", text, "")]


def farewell_rows(face):
    """告别：纯台词 + **末尾一条 Leave**。

    Leave 是这一类唯一让客人真的离开的开关——去掉它，他会一直站在「待告别」占着客房。
    这也是全表唯一一条「由内容决定业务推进」的事件，所以只能放在最后一步：
    它一执行访客就离场、演员开始走向门口，后面还有台词的话会读成「人都走了还在说话」。
    """
    return [
        ("visitor", face, 1, "", "", "Line", "时候不早了，我也该走啦。", ""),
        ("visitor", "", 2, "", "", "Line", "谢谢你这几天的照顾——下次路过，我还来。", ""),
        ("", "", 3, "", "", "Action", "Leave", ""),
    ]


FEEDBACKS = [
    ("feedbackDisappointed", 40001, "算了……可能是我要求太多了。", "服务超时后说的话"),
    ("feedbackPlain", 40101, "嗯，还行吧，谢谢你。", "只有小游戏类会走到"),
    ("feedbackFine", 40201, "挺好的，比我原本想的还妥帖些。", "只有小游戏类会走到"),
    ("feedbackPerfect", 40301, "太完美了，我会记住这地方的。", "条件类交付成功 / 小游戏满分"),
]
SMALL_TALKS = [
    "这屋子比我想的舒服多了。",
    "难得能这么安静地待一会儿。",
    "下次路过，我还来。",
]


def build_all(needs, races, portraits):
    """返回 (第二页内容组, 第一页挂载行)。组 ID 的个位段留给种族下标。

    ID 分段（只是约定，导入器不强制）：
        1xxxx 初次见面   2xxxx 等待接待   3xxxx 需求对话
        4xxxx 需求反馈   5xxxx 闲聊       6xxxx 告别
    每一段里再按「内容序号 × 100 + 种族下标」排开，加种族/加需求都不会撞号。
    """
    content = []
    bindings = []

    for r, race in enumerate(races):
        face = default_portrait_of(race, portraits)

        gid = 10001 + r
        content.append((gid, first_meeting_rows(face)))
        bindings.append((gid, race, "", "firstMeeting", "", "开场：打招呼 + 接待/拒绝分支"))

        gid = 20001 + r
        content.append((gid, waiting_rows(face)))
        bindings.append((gid, race, "", "waitingReception", "",
                         "已经打过招呼、再点他时说的话（同样要带接待/拒绝）"))

        for n, need in enumerate(needs):
            minigame = "电路" in need or "minigame" in need.lower()
            gid = 30001 + n * 100 + r
            content.append((gid, need_rows(face, minigame)))
            bindings.append((gid, race, need, "needTalk", "",
                             "%s需求「%s」的说辞与交付分支" % ("小游戏类" if minigame else "条件类", need)))

        for category, base, text, note in FEEDBACKS:
            gid = base + r
            content.append((gid, one_liner(face, text)))
            bindings.append((gid, race, "", category, "", note))

        for k, text in enumerate(SMALL_TALKS):
            gid = 50001 + k * 100 + r
            content.append((gid, one_liner(face, text)))
            bindings.append((gid, race, "", "smallTalk", "", ""))

        gid = 60001 + r
        content.append((gid, farewell_rows(face)))
        bindings.append((gid, race, "", "farewell", "",
                         "停留到点后点他才播；末尾那条 Leave 才是让他离开的开关"))

    content.sort(key=lambda item: item[0])
    bindings.sort(key=lambda item: item[0])
    return content, bindings


# ── 写表 ────────────────────────────────────────────────────────────────────

def write_header(ws, header, fields, widths):
    ws.append(header)
    ws.append(fields)
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = HEAD_FILL
        ws.cell(row=2, column=col).font = FIELD_FONT
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]
    ws.freeze_panes = "A3"


def add_list_validation(ws, column_letter, formula, last_row=2000, prompt=""):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    if prompt:
        dv.promptTitle = "可选值"
        dv.prompt = prompt
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add("%s3:%s%d" % (column_letter, column_letter, last_row))


def build_workbook(needs, races, portraits):
    wb = openpyxl.Workbook()
    content_groups, bindings = build_all(needs, races, portraits)

    # ── 第一页：对话组 → 池 ──
    ws_group = wb.active
    ws_group.title = GROUP_SHEET
    write_header(ws_group, GROUP_HEADER, GROUP_FIELDS, GROUP_WIDTH)
    for row in bindings:
        ws_group.append(list(row))

    # ── 第二页：对话内容 ──
    ws_content = wb.create_sheet(CONTENT_SHEET)
    write_header(ws_content, CONTENT_HEADER, CONTENT_FIELDS, CONTENT_WIDTH)
    for gid, rows in content_groups:
        first = True
        for speaker, portrait, step, option, sub, kind, text, cond in rows:
            ws_content.append([gid if first else "", speaker, portrait, step, option, sub, kind, text, cond])
            first = False
    for row in ws_content.iter_rows(min_row=3, min_col=8, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # ── 第三页：参考 ──
    ws_ref = wb.create_sheet(REF_SHEET)
    blocks = [
        ("所属对话池", CATEGORIES),
        ("类型", KINDS),
        ("说话人", SPEAKERS),
        ("立绘ID", [(p, "") for p in portraits]),
        ("条件函数", CONDITIONS),
        ("事件函数", ACTIONS),
        ("需求ID", [(n, "") for n in needs]),
        ("种族id", [(r, "") for r in races]),
    ]
    col = 1
    spans = {}
    for title, items in blocks:
        ws_ref.cell(row=1, column=col, value=title).font = Font(bold=True)
        ws_ref.cell(row=1, column=col).fill = REF_FILL
        ws_ref.cell(row=1, column=col + 1, value="说明").font = Font(bold=True)
        ws_ref.cell(row=1, column=col + 1).fill = REF_FILL
        for i, (key, desc) in enumerate(items):
            ws_ref.cell(row=2 + i, column=col, value=key)
            ws_ref.cell(row=2 + i, column=col + 1, value=desc)
        ws_ref.column_dimensions[get_column_letter(col)].width = 24
        ws_ref.column_dimensions[get_column_letter(col + 1)].width = 52
        spans[title] = (get_column_letter(col), 2, 1 + max(len(items), 1))
        col += 3
    ws_ref.freeze_panes = "A2"

    def ref(title, pad=0):
        letter, top, bottom = spans[title]
        return "=%s!$%s$%d:$%s$%d" % (REF_SHEET, letter, top, letter, bottom + pad)

    # ── 下拉 ──
    add_list_validation(ws_group, "D", ref("所属对话池"), prompt="九个触发分类，见「参考」页")
    add_list_validation(ws_group, "C", ref("需求ID", pad=40),
                        prompt="NeedDef 的资产名。needTalk 必填；四档反馈与告别选填；其余分类留空")
    add_list_validation(ws_group, "B", ref("种族id", pad=20),
                        prompt="**只能填一个 raceId**：一个对话组只属于一个种族，不再有「通用」与 / 多选")
    add_list_validation(ws_content, "G", ref("类型"))
    add_list_validation(ws_content, "B", ref("说话人"))
    add_list_validation(ws_content, "C", ref("立绘ID", pad=200),
                        prompt="Excel/立绘表.xlsx 里的立绘ID。**留空 = 沿用上一句**；组内首句留空则用种族默认立绘")
    return wb


def main():
    force = "--force" in sys.argv
    if os.path.exists(EXCEL_PATH) and not force:
        print("[SKIP] %s already exists. Pass --force to rebuild (this DISCARDS existing lines)."
              % os.path.abspath(EXCEL_PATH))
        return

    needs = scan_needs()
    races = scan_races()
    portraits = scan_portraits()
    if not needs:
        print("[WARN] no NeedDef assets found under %s" % NEED_DIR)
    if not races:
        print("[WARN] no VisitorRaceDef assets found under %s" % RACE_DIR)
    if not portraits:
        print("[WARN] no portrait ids found in %s "
              "(run make_portrait_template.py first)" % PORTRAIT_EXCEL_PATH)

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    build_workbook(needs, races, portraits).save(EXCEL_PATH)
    print("[OK] wrote %s" % os.path.abspath(EXCEL_PATH))
    print("     needs:     %s" % (", ".join(needs) or "(none)"))
    print("     races:     %s" % (", ".join(races) or "(none)"))
    print("     portraits: %s" % (", ".join(portraits) or "(none)"))
    print("     groups:    %d (每个种族一套，不再有「通用」)" % (len(races) * (2 + len(needs) + 4 + len(SMALL_TALKS))))


if __name__ == "__main__":
    main()
