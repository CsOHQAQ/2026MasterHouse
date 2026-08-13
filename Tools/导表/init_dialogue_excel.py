# -*- coding: utf-8 -*-
"""从现有对话 ScriptableObject 资产生成 Excel/对话表.xlsx 初始模板（完整版）。
只需运行一次；之后用 export_dialogue.py 正向导出。

表结构：
  Sheet 1 对话内容 - 对话组的步骤/台词/分支选项
  Sheet 2 对话池   - 哪个对话组在哪个触发时机出现
"""
import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(__file__)
ROOT = os.path.join(SCRIPT_DIR, "..", "..")
BASE = os.path.join(ROOT, "Assets", "GameData", "Dialogue")
EXCEL_OUT = os.path.join(ROOT, "Excel", "对话表.xlsx")

SPEAKER_MAP = {"0": "访客", "1": "玩家", "2": "旁白"}
EMOTION_MAP = {"0": "平静", "1": "高兴", "2": "困惑", "3": "失望", "4": "惊讶"}
ALL_RACES = ["crow", "fox", "hedgehog", "rabbit"]
TRIGGERS = [
    "firstMeeting", "serviceStart", "serviceCheck", "rejected",
    "doneMismatch", "donePlain", "doneSatisfied", "donePerfect", "wanderChat",
]
PAT_UNICODE = re.compile(r"\\u([0-9a-fA-F]{4})")


def decode(raw):
    raw = raw.strip()
    if raw.startswith('"') and raw.endswith('"'):
        raw = raw[1:-1]
    raw = PAT_UNICODE.sub(lambda m: chr(int(m.group(1), 16)), raw)
    return raw


# ─── GUID map ──────────────────────────────────────────────────────────────
def build_guid_map():
    guid_to_path = {}
    for root, dirs, files in os.walk(BASE):
        for f in files:
            if f.endswith(".asset.meta"):
                meta_path = os.path.join(root, f)
                with open(meta_path, encoding="utf-8") as mf:
                    for line in mf:
                        m = re.match(r"^guid:\s*(\S+)", line.strip())
                        if m:
                            guid_to_path[m.group(1)] = meta_path[:-5]
                            break
    return guid_to_path


# ─── Parse group asset ─────────────────────────────────────────────────────
def parse_group(path):
    with open(path, encoding="utf-8") as f:
        lines = f.readlines()

    result = {"id": "", "note": "", "steps": []}
    for line in lines:
        s = line.strip()
        if s.startswith("id: "):
            result["id"] = decode(s[4:])
        elif s.startswith("note: "):
            result["note"] = decode(s[6:])

    rid_to_class = {}
    for i, line in enumerate(lines):
        if line.strip() == "references:":
            cur_rid = None
            for l2 in lines[i:]:
                m = re.match(r"\s+- rid:\s*(\d+)", l2)
                if m:
                    cur_rid = m.group(1)
                if cur_rid:
                    m2 = re.match(r"\s+type:\s*\{class:\s*(\w+)", l2)
                    if m2:
                        rid_to_class[cur_rid] = m2.group(1)
            break

    steps_start = None
    for i, line in enumerate(lines):
        if line.strip() == "steps:":
            steps_start = i
            break
    if steps_start is None:
        return result

    current_step = None
    in_line_block = False
    in_options = False
    current_option = None
    in_opt_actions = False

    for line in lines[steps_start + 1:]:
        s = line.strip()
        if s == "references:":
            break

        if re.match(r"\s*- kind:", line):
            if current_option is not None and in_options and current_step is not None:
                current_step["options"].append(current_option)
                current_option = None
            if current_step is not None:
                result["steps"].append(current_step)
            kind_val = s.split(":")[1].strip()
            current_step = {"kind": kind_val, "line": {}, "options": [], "actions": []}
            in_line_block = False
            in_options = False
            in_opt_actions = False
            current_option = None
            continue

        if current_step is None:
            continue

        if s == "options:":
            in_options = True
            in_line_block = False
            continue

        if re.match(r"\s+line:", line) and not in_options:
            in_line_block = True
            continue

        if in_line_block and not in_options:
            if re.match(r"\s+speaker:", line):
                current_step["line"]["speaker"] = s.split(":")[1].strip()
            elif re.match(r"\s+text:", line):
                current_step["line"]["text"] = decode(s[5:])
            elif re.match(r"\s+emotion:", line):
                current_step["line"]["emotion"] = s.split(":")[1].strip()
            continue

        if in_options:
            if re.match(r"\s+- conditions:", line):
                if current_option is not None:
                    current_step["options"].append(current_option)
                current_option = {"text": "", "action": "", "action_param": "",
                                  "jump": "", "jump_group": "", "rids": []}
                in_opt_actions = False
                continue
            if current_option is not None:
                if re.match(r"\s+text:", line):
                    current_option["text"] = decode(line.split(":", 1)[1].strip())
                elif re.match(r"\s+actions:", line):
                    in_opt_actions = True
                elif re.match(r"\s+next:", line):
                    next_val = s.split(":")[1].strip()
                    current_option["jump"] = {"0": "继续", "1": "跳到组", "2": "结束"}.get(next_val, "结束")
                elif in_opt_actions:
                    m_rid = re.match(r"\s+- rid:\s*(\d+)", line)
                    if m_rid:
                        current_option["rids"].append(m_rid.group(1))
                    elif not re.match(r"\s+- ", line) and ":" in line:
                        in_opt_actions = False

    if current_option is not None and in_options and current_step is not None:
        current_step["options"].append(current_option)
    if current_step is not None:
        result["steps"].append(current_step)

    for step in result["steps"]:
        for opt in step.get("options", []):
            for rid in opt.get("rids", []):
                cls = rid_to_class.get(rid, "")
                if "Accept" in cls:
                    opt["action"] = "接待"
                elif "Reject" in cls:
                    opt["action"] = "拒绝"
                elif "CompleteNeed" in cls:
                    opt["action"] = "完成需求"
                elif "StartMinigame" in cls:
                    opt["action"] = "小游戏"
                elif "AddCurrency" in cls:
                    opt["action"] = "货币"
                elif "AddReputation" in cls:
                    opt["action"] = "声望"
                elif "Log" in cls:
                    opt["action"] = "日志"

    return result


# ─── Parse pool asset ──────────────────────────────────────────────────────
def parse_pool(path):
    with open(path, encoding="utf-8") as f:
        content = f.read()

    result = {}
    for trig in TRIGGERS:
        entries = []
        m = re.search(r"  " + trig + r":\n((?:  - .*\n|    .*\n)*)", content)
        if m:
            block = m.group(1)
            for em in re.finditer(r"- group: \{fileID: \d+, guid: ([a-f0-9]+)", block):
                guid = em.group(1)
                pos = em.end()
                wm = re.search(r"weight: (\d+)", block[pos: pos + 100])
                weight = int(wm.group(1)) if wm else 1
                entries.append((guid, weight))
        result[trig] = entries

    return result


# ─── Excel styles ──────────────────────────────────────────────────────────
FONT_NAME = "微软雅黑"
H_FILL = PatternFill("solid", fgColor="2F5496")
H_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True)
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(name=FONT_NAME, color="7F6000", italic=True, size=9)
NOTE_ALIGN = Alignment(vertical="top", wrap_text=True)
DATA_FONT = Font(name=FONT_NAME)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="AAAAAA")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_header(ws, cols):
    for i, (header, width) in enumerate(cols, 1):
        c = ws.cell(1, i, header)
        c.fill = H_FILL
        c.font = H_FONT
        c.alignment = H_ALIGN
        c.border = BD
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22


def set_note_row(ws, notes, row=2):
    for i, note in enumerate(notes, 1):
        c = ws.cell(row, i, note)
        c.fill = NOTE_FILL
        c.font = NOTE_FONT
        c.alignment = NOTE_ALIGN
        c.border = BD
    ws.row_dimensions[row].height = 52


def style_data(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row, c)
        cell.border = BD
        cell.alignment = DATA_ALIGN
        if cell.font is None or cell.font.name != FONT_NAME:
            cell.font = DATA_FONT


# ─── Main ──────────────────────────────────────────────────────────────────
def main():
    guid_map = build_guid_map()

    group_data = {}
    for root, dirs, files in os.walk(BASE):
        for f in sorted(files):
            if not f.endswith(".asset") or f.endswith(".meta") or f.startswith("Pool_"):
                continue
            path = os.path.join(root, f)
            rel = path[len(BASE):].replace("\\", "/").lstrip("/")
            folder = rel.rsplit("/", 1)[0] if "/" in rel else "通用"
            data = parse_group(path)
            name = f[:-6]
            if not data["id"]:
                data["id"] = name
            group_data[name] = {"folder": folder, "note": data["note"],
                                "steps": data["steps"], "name": name}

    pool_map = defaultdict(lambda: defaultdict(dict))
    for race in ALL_RACES:
        pool_path = os.path.join(BASE, f"Pool_{race}.asset")
        if not os.path.exists(pool_path):
            continue
        pool = parse_pool(pool_path)
        for trig, entries in pool.items():
            for guid, weight in entries:
                asset_path = guid_map.get(guid, "")
                if not asset_path:
                    continue
                gname = os.path.basename(asset_path)[:-6]
                rel2 = asset_path[len(BASE):].replace("\\", "/").lstrip("/")
                folder2 = rel2.rsplit("/", 1)[0] if "/" in rel2 else "通用"
                pool_map[(gname, folder2, trig)][race] = weight

    wb = openpyxl.Workbook()

    # ── Sheet 1: 对话内容 ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "对话内容"
    ws1.freeze_panes = "A3"

    CONTENT_COLS = [
        ("对话组ID", 20), ("备注", 18), ("文件夹", 9), ("步骤", 5),
        ("类型", 7), ("说话人", 7), ("表情", 7), ("文本", 40),
        ("动作", 9), ("动作参数", 10), ("跳转", 8), ("跳转目标组", 16), ("选项条件", 22),
    ]
    set_header(ws1, CONTENT_COLS)
    set_note_row(ws1, [
        "【必填】对话组资产名，同组所有行填相同ID",
        "可选备注，仅编辑器可见不进游戏",
        "资产子目录，默认【通用】",
        "【必填】步骤序号1起；同编号多行=分支",
        "台词 / 选项 / 事件\n台词=显示文字\n选项=玩家可选\n事件=触发效果不显示",
        "台词填：访客/玩家/旁白\n选项/事件留空",
        "仅访客台词：平静/高兴/困惑/失望/惊讶",
        "台词正文或选项文字\n占位符：{访客名}{需求}{物品名}",
        "接待/拒绝/完成需求/小游戏/货币/声望/日志\n台词步骤留空",
        "完成需求填满意度档位（不填=完美）：不对味/一般/满意/完美\n货币/声望填整数（正数加负数减）\n日志填消息文本",
        "仅选项行：结束（默认）/继续/跳到组",
        "跳转=跳到组时填目标对话组ID",
        "仅选项/对话组进入条件，多条用;分隔\n格式：天数>=N/货币>=N/声望>=N\n种族:race_id/满意度>=满意\n访客状态:枚举名/有空房/房间有家具",
    ])

    row = 3
    for name in sorted(group_data):
        gdata = group_data[name]
        steps = gdata["steps"]
        first = True

        if not steps:
            vals = [gdata["name"], gdata["note"], gdata["folder"]] + [""] * (len(CONTENT_COLS) - 3)
            for c, v in enumerate(vals, 1):
                ws1.cell(row, c, v)
            style_data(ws1, row, len(CONTENT_COLS))
            row += 1
            continue

        for si, step in enumerate(steps):
            kind = step["kind"]
            if kind == "0":  # Line
                line = step.get("line", {})
                spk_key = line.get("speaker", "0")
                speaker = SPEAKER_MAP.get(spk_key, "访客")
                emotion = EMOTION_MAP.get(line.get("emotion", "0"), "平静") if speaker == "访客" else ""
                vals = [
                    gdata["name"] if first else "",
                    gdata["note"] if first else "",
                    gdata["folder"] if first else "",
                    str(si + 1), "台词", speaker, emotion,
                    line.get("text", ""), "", "", "", "", "",
                ]
                for c, v in enumerate(vals, 1):
                    ws1.cell(row, c, v)
                style_data(ws1, row, len(CONTENT_COLS))
                first = False
                row += 1
            elif kind == "2":  # Branch
                for opt in step.get("options", []):
                    jump = opt.get("jump", "结束") or "结束"
                    vals = [
                        gdata["name"] if first else "",
                        gdata["note"] if first else "",
                        gdata["folder"] if first else "",
                        str(si + 1), "选项", "", "",
                        opt.get("text", ""),
                        opt.get("action", ""), opt.get("action_param", ""),
                        jump,
                        opt.get("jump_group", "") if jump == "跳到组" else "",
                        "",
                    ]
                    for c, v in enumerate(vals, 1):
                        ws1.cell(row, c, v)
                    style_data(ws1, row, len(CONTENT_COLS))
                    first = False
                    row += 1

    # ── Sheet 2: 对话池 ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("对话池")
    ws2.freeze_panes = "A3"

    POOL_COLS = [
        ("对话组ID", 22), ("文件夹", 9), ("种族", 11), ("触发分类", 17), ("权重", 5), ("进入条件", 22),
    ]
    set_header(ws2, POOL_COLS)
    set_note_row(ws2, [
        "【必填】对话组资产名",
        "资产子目录，不填=通用",
        "【必填】crow/fox/hedgehog/rabbit\n或【通用】=同时加入全4种族",
        "【必填】firstMeeting 初次见面\nserviceStart 开始服务\nserviceCheck 服务中交谈\nrejected 被拒绝\ndoneMismatch 完成·不对味\ndonePlain 完成·一般\ndoneSatisfied 完成·满意\ndonePerfect 完成·完美\nwanderChat 闲逛气泡",
        "抽取权重正整数，不填=1",
        "该对话组参与抽取的前提条件\n格式同选项条件列\n多条用;分隔，留空=无条件",
    ])

    row2 = 3
    for (gname, folder, trig), races in sorted(pool_map.items(), key=lambda x: (x[0][0], x[0][2])):
        all_four = set(races.keys()) == set(ALL_RACES)
        weight = list(races.values())[0]
        if all_four:
            for c, v in enumerate([gname, folder, "通用", trig, weight, ""], 1):
                ws2.cell(row2, c, v)
            style_data(ws2, row2, len(POOL_COLS))
            row2 += 1
        else:
            for race in sorted(races):
                for c, v in enumerate([gname, folder, race, trig, races[race], ""], 1):
                    ws2.cell(row2, c, v)
                style_data(ws2, row2, len(POOL_COLS))
                row2 += 1

    os.makedirs(os.path.dirname(EXCEL_OUT), exist_ok=True)
    wb.save(EXCEL_OUT)
    print(f"[OK] {EXCEL_OUT}")
    print(f"     对话内容: {row - 3} 行")
    print(f"     对话池:   {row2 - 3} 行")


if __name__ == "__main__":
    main()
